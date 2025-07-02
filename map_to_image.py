import json
import yaml

import csv
import numpy as np
import cv2
from openpyxl import load_workbook


from openpyxl import load_workbook
from openpyxl.styles import PatternFill

#Traversabel_grids : #ff9900
#00ffff

grid_types = {

"feeding_grid"      : "00ff00",
"dumping_grid"      : "ffff00",
"queuing_grid"      :"4a86e8",

}

def map2_image(grid_map):
    graw_scale_imag = np.zeros((len(grid_map),len(grid_map[0]),1), dtype=np.uint8)
    
    for row in range(len(grid_map)):
        for coloumn in range(len(grid_map[0])):
            if ',' in grid_map[row][coloumn]:
                graw_scale_imag[row][coloumn] = 255
            else :
                graw_scale_imag[row][coloumn] = 0
    cv2.imwrite('output.png',graw_scale_imag)

def extract_colored_cells(filename, sheet_name=None):
    """
    Load the workbook, scan the specified sheet (or the active one),
    and return a dict mapping each grid name to the list of cell coords.
    """
    wb = load_workbook(filename, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    results = {name: [] for name in grid_types}
    for row in ws.iter_rows():
        for cell in row:
            fill = cell.fill
            if  fill.fill_type == "solid":
                rgb = fill.fgColor.rgb  
                hexcolor = rgb[-6:].lower()
                for grid_name, target in grid_types.items():
                    if hexcolor == target:
                        results[grid_name].append((cell.row - 1,cell.column -1 ))
    for grid_name, coords in results.items():
        coords.sort(key=lambda t: (t[1], t[0]))
        
    return results

def load_csv(path):
    with open(path, newline='') as f:
        return [row for row in csv.reader(f)]
    
if __name__ == '__main__':
    grid_raw = load_csv("/home/sai/projects/lexxpluss/map.csv")
    # convert grid_raw into grid_map[y][x]dr
    map2_image(grid_raw)
    dump_grid_config = extract_colored_cells("/home/sai/projects/lexxpluss/map.xlsx")
    print(dump_grid_config)
    with open('dump_grid_config.json', 'w') as json_file:
        json.dump(dump_grid_config, json_file,indent=2)
    
